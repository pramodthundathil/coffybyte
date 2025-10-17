from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, logout, login
from django.contrib import messages 
from django.contrib.auth.hashers import make_password

from .decorators import store_owner_access
from inventory.models import *
from authentication.models import *
from .forms import MenuForm, TablesForm, StoreAddUserForm
from orders.models import *



def signin(request):
    if request.method == "POST":
        username = request.POST['uname']
        password = request.POST['pswd']
        user1 = authenticate(request, username = username , password = password)
        
        if user1 is not None:
            
            request.session['username'] = username
            request.session['password'] = password
            login(request, user1)
            return redirect('dashboard')
        
        else:
            messages.error(request,'Username or Password Incorrect')
            return redirect('SignIn')
    return render(request,"login.html")






# @store_owner_access   
# def dashboard(request):
#     return render(request, 'index.html')


# views.py
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db.models import Sum, Count, Avg, Q
from django.db.models.functions import TruncDate, TruncMonth, TruncWeek
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import json

from authentication.models import Store, Branch, CustomUser, StoreUser
from inventory.models import Menu, FoodCategory, Tax
from django.db.models.functions import ExtractWeekDay


from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
import json
from datetime import datetime
import csv

@login_required
def refresh_metrics(request):
    """AJAX endpoint to refresh dashboard metrics without full page reload"""
    try:
        store = request.user.store_memberships.all()[0].store
    except (IndexError, AttributeError):
        return JsonResponse({'error': 'No store found'}, status=400)
    
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)
    
    # Get refreshed stats
    today_stats = get_today_stats(store, today)
    comparison_stats = get_comparison_stats(store, today, yesterday)
    
    return JsonResponse({
        'today_revenue': float(today_stats['revenue']),
        'today_orders': today_stats['orders_count'],
        'avg_order_value': float(today_stats['avg_order_value']),
        'revenue_change': comparison_stats['revenue_change'],
        'orders_change': comparison_stats['orders_change'],
        'last_updated': datetime.now().strftime('%H:%M')
    })

@login_required
def export_dashboard_report(request):
    """Export dashboard data as CSV"""
    try:
        store = request.user.store_memberships.all()[0].store
    except (IndexError, AttributeError):
        return JsonResponse({'error': 'No store found'}, status=400)
    
    # Create the HttpResponse object with CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{store.name}_dashboard_report_{datetime.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    
    # Write headers
    writer.writerow(['Date', 'Orders', 'Revenue', 'Avg Order Value'])
    
    # Get last 30 days data
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    daily_data = Order.objects.filter(
        store=store,
        create_date__date__gte=start_date,
        create_date__date__lte=end_date,
        checkout_status=True
    ).annotate(
        date=TruncDate('create_date')
    ).values('date').annotate(
        orders_count=Count('id'),
        revenue=Sum('total_price')
    ).order_by('date')
    
    # Write data rows
    for item in daily_data:
        avg_value = item['revenue'] / item['orders_count'] if item['orders_count'] > 0 else 0
        writer.writerow([
            item['date'].strftime('%Y-%m-%d'),
            item['orders_count'],
            f"${item['revenue']:.2f}",
            f"${avg_value:.2f}"
        ])
    
    return response

@store_owner_access
def dashboard(request):
    """Main dashboard view for store owner"""
    try:
        # Get store from user's membership
        store = request.user.store_memberships.all()[0].store
    except (IndexError, AttributeError):
        return render(request, 'index.html')
    
    # Date filters
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    
    # Basic stats
    context = {
        'store': store,
        'today_stats': get_today_stats(store, today),
        'week_stats': get_week_stats(store, week_start, today),
        'month_stats': get_month_stats(store, month_start, today),
        'comparison_stats': get_comparison_stats(store, today, yesterday),
        'top_items': get_top_selling_items(store, today - timedelta(days=30), today),
        'recent_orders': get_recent_orders(store),
        'payment_breakdown': get_payment_breakdown(store, today),
        'order_status_breakdown': get_order_status_breakdown(store, today),
        'branch_performance': get_branch_performance(store, today),
        'staff_performance': get_staff_performance(store, today),
    }
    
    return render(request, 'dashboard_index.html', context)

def get_today_stats(store, today):
    """Get today's key metrics"""
    today_orders = Order.objects.filter(
        store=store, 
        create_date__date=today,
        checkout_status=True
    )
    
    today_revenue = today_orders.aggregate(
        total=Sum('total_price')
    )['total'] or Decimal('0.00')
    
    today_orders_count = today_orders.count()
    
    # Average order value
    avg_order_value = today_revenue / today_orders_count if today_orders_count > 0 else Decimal('0.00')
    
    # Most sold items today
    popular_items = OrderItem.objects.filter(
        order__store=store,
        order__create_date__date=today,
        order__checkout_status=True
    ).values('menu_item__name').annotate(
        total_quantity=Sum('quantity')
    ).order_by('-total_quantity')[:3]
    
    return {
        'revenue': today_revenue,
        'orders_count': today_orders_count,
        'avg_order_value': avg_order_value,
        'popular_items': popular_items,
    }

def get_week_stats(store, week_start, today):
    """Get this week's metrics"""
    week_orders = Order.objects.filter(
        store=store,
        create_date__date__gte=week_start,
        create_date__date__lte=today,
        checkout_status=True
    )
    
    week_revenue = week_orders.aggregate(
        total=Sum('total_price')
    )['total'] or Decimal('0.00')
    
    return {
        'revenue': week_revenue,
        'orders_count': week_orders.count(),
    }

def get_month_stats(store, month_start, today):
    """Get this month's metrics"""
    month_orders = Order.objects.filter(
        store=store,
        create_date__date__gte=month_start,
        create_date__date__lte=today,
        checkout_status=True
    )
    
    month_revenue = month_orders.aggregate(
        total=Sum('total_price')
    )['total'] or Decimal('0.00')
    
    return {
        'revenue': month_revenue,
        'orders_count': month_orders.count(),
    }

def get_comparison_stats(store, today, yesterday):
    """Compare today vs yesterday"""
    yesterday_orders = Order.objects.filter(
        store=store,
        create_date__date=yesterday,
        checkout_status=True
    )
    
    today_orders = Order.objects.filter(
        store=store,
        create_date__date=today,
        checkout_status=True
    )
    
    yesterday_revenue = yesterday_orders.aggregate(
        total=Sum('total_price')
    )['total'] or Decimal('0.00')
    
    today_revenue = today_orders.aggregate(
        total=Sum('total_price')
    )['total'] or Decimal('0.00')
    
    # Calculate percentage changes
    revenue_change = 0
    orders_change = 0
    
    if yesterday_revenue > 0:
        revenue_change = float(((today_revenue - yesterday_revenue) / yesterday_revenue) * 100)
    
    yesterday_count = yesterday_orders.count()
    today_count = today_orders.count()
    
    if yesterday_count > 0:
        orders_change = ((today_count - yesterday_count) / yesterday_count) * 100
    
    return {
        'revenue_change': round(revenue_change, 1),
        'orders_change': round(orders_change, 1),
    }

def get_top_selling_items(store, start_date, end_date):
    """Get top selling items in date range"""
    return OrderItem.objects.filter(
        order__store=store,
        order__create_date__date__gte=start_date,
        order__create_date__date__lte=end_date,
        order__checkout_status=True
    ).values(
        'menu_item__name',
        'menu_item__price'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum('price')
    ).order_by('-total_quantity')[:10]

def get_recent_orders(store):
    """Get recent orders for the store"""
    return Order.objects.filter(
        store=store
    ).select_related('table', 'user').order_by('-create_date')[:10]

def get_payment_breakdown(store, date):
    """Get payment method breakdown for today"""
    return Checkout.objects.filter(
        order__store=store,
        datetime__date=date
    ).values('payment_method').annotate(
        count=Count('id'),
        total_amount=Sum('total_price')
    ).order_by('-total_amount')

def get_order_status_breakdown(store, date):
    """Get order status breakdown"""
    return Order.objects.filter(
        store=store,
        create_date__date=date
    ).values('status').annotate(
        count=Count('id')
    ).order_by('-count')

def get_branch_performance(store, date):
    """Get branch-wise performance if multiple branches"""
    # Note: Based on your model, orders don't directly link to branches
    # You might need to add branch field to Order model or get it through user
    branches = store.branches.all()
    branch_data = []
    
    for branch in branches:
        # Get orders from users assigned to this branch
        branch_users = branch.branch_users.values_list('user_id', flat=True)
        branch_orders = Order.objects.filter(
            store=store,
            user_id__in=branch_users,
            create_date__date=date,
            checkout_status=True
        )
        
        branch_revenue = branch_orders.aggregate(
            total=Sum('total_price')
        )['total'] or Decimal('0.00')
        
        branch_data.append({
            'name': branch.name,
            'orders_count': branch_orders.count(),
            'revenue': branch_revenue
        })
    
    return branch_data

def get_staff_performance(store, date):
    """Get staff performance data"""
    staff_performance = Order.objects.filter(
        store=store,
        create_date__date=date,
        checkout_status=True,
        user__isnull=False
    ).values(
        'user__first_name',
        'user__last_name'
    ).annotate(
        orders_count=Count('id'),
        total_revenue=Sum('total_price')
    ).order_by('-total_revenue')[:10]
    
    return staff_performance

# AJAX API endpoints for charts
@login_required
def dashboard_chart_data(request, chart_type):
    """API endpoint for chart data"""
    try:
        store = request.user.store_memberships.all()[0].store
    except (IndexError, AttributeError):
        return JsonResponse({'error': 'No store found'}, status=400)
    
    if chart_type == 'revenue_trend':
        return get_revenue_trend_data(request, store)
    elif chart_type == 'category_sales':
        return get_category_sales_data(request, store)
    elif chart_type == 'hourly_orders':
        return get_hourly_orders_data(request, store)
    elif chart_type == 'weekly_comparison':
        return get_weekly_comparison_data(request, store)
    else:
        return JsonResponse({'error': 'Invalid chart type'}, status=400)

def get_revenue_trend_data(request, store):
    """Get revenue trend for last 30 days"""
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    daily_revenue = Order.objects.filter(
        store=store,
        create_date__date__gte=start_date,
        create_date__date__lte=end_date,
        checkout_status=True
    ).annotate(
        date=TruncDate('create_date')
    ).values('date').annotate(
        revenue=Sum('total_price')
    ).order_by('date')
    
    return JsonResponse({
        'labels': [item['date'].strftime('%m-%d') for item in daily_revenue],
        'data': [float(item['revenue']) for item in daily_revenue]
    })

def get_category_sales_data(request, store):
    """Get category-wise sales data"""
    category_sales = OrderItem.objects.filter(
        order__store=store,
        order__create_date__date=timezone.now().date(),
        order__checkout_status=True
    ).values(
        'menu_item__category__name'
    ).annotate(
        total_revenue=Sum('price'),
        total_quantity=Sum('quantity')
    ).order_by('-total_revenue')
    
    return JsonResponse({
        'labels': [item['menu_item__category__name'] for item in category_sales],
        'revenue': [float(item['total_revenue']) for item in category_sales],
        'quantity': [item['total_quantity'] for item in category_sales]
    })

def get_hourly_orders_data(request, store):
    """Get hourly order distribution for today"""
    
    
    hourly_orders = Order.objects.filter(
        store=store,
        create_date__date=timezone.now().date()
    ).annotate(
        hour=ExtractWeekDay('create_date', 'hour')
    ).values('hour').annotate(
        count=Count('id')
    ).order_by('hour')
    
    # Create 24-hour data
    hourly_data = [0] * 24
    for item in hourly_orders:
        hourly_data[item['hour']] = item['count']
    
    return JsonResponse({
        'labels': [f'{i}:00' for i in range(24)],
        'data': hourly_data
    })

def get_weekly_comparison_data(request, store):
    """Get current week vs last week comparison"""
    today = timezone.now().date()
    current_week_start = today - timedelta(days=today.weekday())
    last_week_start = current_week_start - timedelta(days=7)
    last_week_end = current_week_start - timedelta(days=1)
    
    current_week_data = Order.objects.filter(
        store=store,
        create_date__date__gte=current_week_start,
        create_date__date__lte=today,
        checkout_status=True
    ).annotate(
        day=ExtractWeekDay('create_date', 'weekday')
    ).values('day').annotate(
        revenue=Sum('total_price')
    ).order_by('day')
    
    last_week_data = Order.objects.filter(
        store=store,
        create_date__date__gte=last_week_start,
        create_date__date__lte=last_week_end,
        checkout_status=True
    ).annotate(
        day=ExtractWeekDay('create_date', 'weekday')
    ).values('day').annotate(
        revenue=Sum('total_price')
    ).order_by('day')
    
    # Convert to arrays
    current_week_revenue = [0.0] * 7
    last_week_revenue = [0.0] * 7
    
    for item in current_week_data:
        current_week_revenue[item['day'] - 1] = float(item['revenue'])
    
    for item in last_week_data:
        last_week_revenue[item['day'] - 1] = float(item['revenue'])
    
    return JsonResponse({
        'labels': ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
        'current_week': current_week_revenue,
        'last_week': last_week_revenue
    })

def SignOut(request):
    logout(request)
    return redirect("SignIn")




# product functions 

@store_owner_access 
def Add_Category(request):
    if request.method == "POST":
        pic = request.FILES['pic']
        cname = request.POST['cname']
        foodcategory = FoodCategory.objects.create(image = pic, name= cname, store = request.user.store_memberships.all()[0].store )
        foodcategory.save()
        messages.success(request,"Food Category Addedd...")
        return redirect("List_Category")
    
    return render(request,'fooditems/add-category.html')

@store_owner_access 
def List_Category(request):
    store = request.user.store_memberships.all()[0].store
    food_category = FoodCategory.objects.filter(store = store)
   
    context = {
        "food_category":food_category
    }
    return render(request,'fooditems/list-category.html',context)


@store_owner_access 
def EditCategory(request,pk):

    cat = get_object_or_404(FoodCategory, id=pk)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        image = request.FILES.get('image')
        
        if name:
            cat.name = name
        
        if image:
            cat.image = image
        
        try:
            cat.save()
            messages.success(request, 'Category updated successfully')
            return redirect('EditCategory', pk = pk)  
        except Exception as e:
            messages.error(request, f'Error updating category: {e}')
    
    context = {
        'cat': cat,
    }
    return render(request,'fooditems/edit-category.html',context)


@store_owner_access 
def DeleteCategory(request,pk):
    cat = FoodCategory.objects.get(id = pk)
    cat.delete()
    messages.success(request,'Food Category Deleted')
    return redirect('List_Category')



# finished basic testing of category finish category functions ---------------------- 
# finished basic testing of category finish category functions ----------------------
# 
#  tax need to find the testing 
@store_owner_access 
def AddTax(request):
    if request.method == "POST":
        name = request.POST.get('name')
        tax_rate = request.POST.get('tax')
        tax = Tax.objects.create(tax_name = name,tax_percentage = tax_rate )
        tax.save()
        messages.success(request,'Tax Value Added Success')
        return redirect("ListTax")
    return render(request,"add-tax-slab.html")

@store_owner_access 
def ListTax(request):
    tax = Tax.objects.all()

    context = {
        "tax":tax
    }
    return render(request,"list-tax.html",context)


@store_owner_access 
def Add_Product(request):
    if request.method == "POST":
        form = MenuForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            menu_item = form.save(commit=False)
            menu_item.store = request.user.store_memberships.all()[0].store
            menu_item.save()
            messages.success(request,"menu Created")
            return redirect("List_Product")  # replace with your menu list view name
        else:
            messages.error(request,f"Something wrong - Error {form.errors.as_text}")
            return redirect("Add_Product")
    form = MenuForm(user=request.user)
        
    context = {
        "form":form
    }
    return render(request,'fooditems/add-product.html',context)

@store_owner_access 
def List_Product(request):
    menu = Menu.objects.filter(store = request.user.store_memberships.all()[0].store)

    context = {
        "menu":menu,
    }
    return render(request,'fooditems/list-product.html',context)

@store_owner_access 
def EditProduct(request, pk):
    menu_item = get_object_or_404(Menu, id = pk)
    if request.method == "POST":
        form = MenuForm(request.POST, request.FILES, user=request.user,instance = menu_item )
        if form.is_valid():
            menu_item = form.save()
            messages.success(request, 'Menu updated successfully')
            return redirect("List_Product")  # replace with your menu list view name
        else:
            messages.error(request,f"Something wrong - Error {form.errors.as_text}")
            return redirect("EditProduct")
    form = MenuForm(user=request.user, instance = menu_item )
        
    
    context = {
        'form':form
    }
    return render(request, "fooditems/edit-product.html", context)

@store_owner_access 
def DeleteProduct(request,pk):
    menu  = Menu.objects.get(id = pk)
    if menu.status == False:
        menu.status = True
    else:
        menu.status = False
    menu.save()
    messages.info(request,"Product Deleted....")
    return redirect("List_Product")


@store_owner_access 
def Add_Table(request):
    if request.method == "POST":
        tnum = request.POST['tnum']
        seats = request.POST['seats']
        if Tables.objects.filter(Table_number = tnum).exists():
            messages.error(request,"Table Already Exists...")
            return redirect("Add_Table")
        else:
            table = Tables.objects.create(Table_number = tnum, Number_of_Seats = seats)
            table.save()
            messages.success(request,"Table added Success...")
            return redirect("List_Table")

    return render(request,"utils/add-table.html")

@store_owner_access 
def List_Table(request):
    table = Tables.objects.all()
    context = {
        "table":table
    }
    return render(request,"utils/list-table.html",context)


@store_owner_access
def edit_table(request, pk):
    table = get_object_or_404(Tables, id=pk)
    form = TablesForm(instance=table)
    if request.method == "POST":
        form = TablesForm(request.POST, instance=table)
        if form.is_valid():
            form.save()
            messages.success(request, "Table updated successfully.")
            return redirect('List_Table')
        else:
            messages.error(request, "Failed to update table.")
    context = {
        'form': form,
        'table': table,
    }
    return render(request, 'utils/edit_table.html', context)

@store_owner_access 
def Delete_Table(request,pk):
    Tables.objects.get(id = pk).delete()
    messages.success(request,"Table deleted success....")
    return redirect("List_Table")


# user management 

# Get all users of a store where a given user is store_owner
def get_store_users(owner_user):
    try:
        # Find the store(s) where this user is the store_owner
        store_user = StoreUser.objects.get(user=owner_user, role="store_owner")
        store = store_user.store

        # Get all users of the same store
        users = CustomUser.objects.filter(store_memberships__store=store).distinct()
        return users
    except StoreUser.DoesNotExist:
        return CustomUser.objects.none()


@store_owner_access 
def ListUser(request):
    contacts = get_store_users(request.user)

    context = {
        'contacts':contacts
    }

    return render(request,"users/user-list.html",context)


@store_owner_access 
def AddUser(request):
    store = request.user.store_memberships.all()[0].store
    if request.method == "POST":
        form = StoreAddUserForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.cleaned_data
            user, created = CustomUser.objects.get_or_create(
                email=data["email"],
                defaults={
                    "first_name": data["first_name"],
                    "last_name": data["last_name"],
                    "phone": data["phone"],
                    "password": make_password(data["password"]),
                },
            )
            # Assign user to store
            StoreUser.objects.get_or_create(
                store=store,
                user=user,
                defaults={"role": data["role"], "assigned_by": request.user}
            )
            messages.success(request, "User added successfully!")
            return redirect("ListUser", store_id=store.id)
    else:
        form = StoreAddUserForm()
    return render(request,"users/user-add.html",{"form": form, "store": store})

@store_owner_access 
def DeleteUser(request,pk):
    CustomUser.objects.get(id = pk).delete()
    messages.success(request,"User Data Deleted.....")
    return redirect("ListUser")


# views.py
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.http import JsonResponse
from datetime import datetime, timedelta
from orders.models import Order, OrderItem, Checkout
from django.contrib import messages




@store_owner_access 
def list_sale(request):
    # Get user's store
    store = request.user.store_memberships.all()[0].store
    
    # Get filter parameters
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    payment_status_filter = request.GET.get('payment_status', '')
    order_method_filter = request.GET.get('order_method', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base queryset - only completed orders (sales)
    orders = Order.objects.filter(
        store=store,
        checkout_status=True  # Only orders that have been checked out
    ).select_related('table', 'user').prefetch_related('items__menu_item')
    
    # Apply filters
    if search_query:
        orders = orders.filter(
            Q(token__icontains=search_query) |
            Q(id__icontains=search_query) |
            Q(table__Table_number__icontains=search_query)
        )
    
    if status_filter:
        orders = orders.filter(status=status_filter)
    
    if payment_status_filter:
        orders = orders.filter(payment_status=payment_status_filter)
    
    if order_method_filter:
        orders = orders.filter(order_method=order_method_filter)
    
    if date_from:
        orders = orders.filter(create_date__date__gte=date_from)
    
    if date_to:
        orders = orders.filter(create_date__date__lte=date_to)
    
    # Order by latest first
    orders = orders.order_by('-create_date')
    
    # Calculate summary statistics
    total_sales = orders.aggregate(
        total_amount=Sum('total_price'),
        total_orders=Count('id'),
        total_tax=Sum('total_tax')
    )
    
    # Pagination
    paginator = Paginator(orders, 25)  # Show 25 orders per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get filter choices for dropdowns
    status_choices = Order.status_options
    payment_status_choices = [
        ("Pending", "Pending"), 
        ("Paid", "Paid"), 
        ("Partial", "Partial")
    ]
    order_method_choices = Order.ORDER_METHOD_CHOICES
    
    context = {
        'orders': page_obj,
        'total_sales': total_sales,
        'search_query': search_query,
        'status_filter': status_filter,
        'payment_status_filter': payment_status_filter,
        'order_method_filter': order_method_filter,
        'date_from': date_from,
        'date_to': date_to,
        'status_choices': status_choices,
        'payment_status_choices': payment_status_choices,
        'order_method_choices': order_method_choices,
        'store': store,
    }
    
    return render(request, 'sales/sales.html', context)


@store_owner_access
def sale_detail(request, order_id):
    """View individual sale details"""
    store = request.user.store_memberships.all()[0].store
    
    # Get the order with all related data
    order = get_object_or_404(
        Order.objects.select_related('table', 'user')
                    ,
        id=order_id,
        store=store,
        checkout_status=True
    )
    
    # Get checkout details if exists
    checkout = None
    try:
        checkout = order.checkout
    except:
        pass
    
    # Get only items that were in checkout (not saved for later)
    order_items = order.items.filter(is_saved_for_later=False)
    
    # Calculate item-wise totals
    items_with_totals = []
    for item in order_items:
        addon_total = sum(addon.price for addon in item.add_ons.all())
        item_subtotal = (item.price + addon_total) * item.quantity
        
        # Calculate tax for this item
        tax_amount = 0
        for tax in item.tax.all():
            tax_amount += (item_subtotal * tax.tax_percentage) / 100
        
        items_with_totals.append({
            'item': item,
            'addon_total': addon_total,
            'subtotal': item_subtotal,
            'tax_amount': tax_amount,
            'total_with_tax': item_subtotal + tax_amount
        })
    
    context = {
        'order': order,
        'checkout': checkout,
        'items_with_totals': items_with_totals,
        'store': store,
    }
    
    return render(request, 'sales/sale_detail.html', context)


@store_owner_access
def update_order_status(request, order_id):
    """AJAX view to update order status"""
    if request.method == 'POST':
        store = request.user.store_memberships.all()[0].store
        order = get_object_or_404(Order, id=order_id, store=store)
        
        new_status = request.POST.get('status')
        if new_status in dict(Order.status_options):
            order.status = new_status
            order.save()
            return JsonResponse({
                'success': True,
                'message': f'Order status updated to {new_status}'
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request'})


@store_owner_access
def sales_analytics(request):
    """View for sales analytics and reports"""
    store = request.user.store_memberships.all()[0].store
    
    # Get date range (default to last 30 days)
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=30)
    
    if request.GET.get('start_date'):
        start_date = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
    if request.GET.get('end_date'):
        end_date = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
    
    # Get sales data
    sales_data = Order.objects.filter(
        store=store,
        checkout_status=True,
        create_date__date__range=[start_date, end_date]
    ).aggregate(
        total_revenue=Sum('total_price'),
        total_orders=Count('id'),
        total_tax_collected=Sum('total_tax'),
        avg_order_value=Sum('total_price') / Count('id') if Count('id') > 0 else 0
    )
    
    # Sales by method
    sales_by_method = Order.objects.filter(
        store=store,
        checkout_status=True,
        create_date__date__range=[start_date, end_date]
    ).values('order_method').annotate(
        count=Count('id'),
        revenue=Sum('total_price')
    ).order_by('-revenue')
    
    # Top selling items
    top_items = OrderItem.objects.filter(
        order__store=store,
        order__checkout_status=True,
        order__create_date__date__range=[start_date, end_date],
        is_saved_for_later=False
    ).values(
        'menu_item__name',
        'menu_item__food_category__name'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum('price') * Sum('quantity')
    ).order_by('-total_quantity')[:10]
    
    context = {
        'sales_data': sales_data,
        'sales_by_method': sales_by_method,
        'top_items': top_items,
        'start_date': start_date,
        'end_date': end_date,
        'store': store,
    }
    
    return render(request, 'sales/analytics.html', context)





# reports 

from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q, F, Avg
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import io
from inventory.models import Menu, Tax, FoodCategory
from authentication.models import Store


@login_required
def reports_dashboard(request):
    """Main reports dashboard view"""
    try:
        # Get user's store
        store = request.user.store_memberships.all()[0].store
        
        # Get today's date
        today = timezone.now().date()
        
        # Calculate quick statistics
        today_orders = Order.objects.filter(store=store, create_date__date=today)
        today_orders_count = today_orders.count()
        
        today_revenue = today_orders.filter(
            checkout_status=True, 
            payment_status='Paid'
        ).aggregate(total=Sum('total_price'))['total'] or Decimal('0.00')
        
        pending_orders_count = Order.objects.filter(
            store=store, 
            status__in=['Pending', 'In Kitchen', 'In Progress']
        ).count()
        
        # This month's revenue
        first_day_of_month = today.replace(day=1)
        month_revenue = Order.objects.filter(
            store=store,
            create_date__date__gte=first_day_of_month,
            checkout_status=True,
            payment_status='Paid'
        ).aggregate(total=Sum('total_price'))['total'] or Decimal('0.00')
        
        context = {
            'today_orders_count': today_orders_count,
            'today_revenue': today_revenue,
            'pending_orders_count': pending_orders_count,
            'month_revenue': month_revenue,
        }
        
        return render(request, 'reports/dashboard.html', context)
        
    except IndexError:
        return render(request, 'error.html', {'message': 'No store assigned to user'})


@login_required
def generate_daybook(request):
    """Generate Day Book Report (Excel/PDF)"""
    if request.method == 'POST':
        try:
            store = request.user.store_memberships.all()[0].store
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            format_type = request.POST.get('format', 'excel')
            
            # Parse dates
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            # Get orders for the date range
            orders = Order.objects.filter(
                store=store,
                create_date__date__gte=start_date,
                create_date__date__lte=end_date,
                checkout_status=True
            ).select_related('table', 'user').prefetch_related('items__menu_item')
            
            if format_type == 'excel':
                return generate_daybook_excel(orders, start_date, end_date, store)
            else:
                return generate_daybook_pdf(orders, start_date, end_date, store)
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)


def generate_daybook_excel(orders, start_date, end_date, store):
    """Generate Excel Day Book Report"""
    # Create a new workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Day Book Report"
    
    # Set up styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=16)
    
    # Add title
    ws['A1'] = f"{store.name} - Day Book Report"
    ws['A1'].font = title_font
    ws['A2'] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    
    # Merge cells for title
    ws.merge_cells('A1:J1')
    ws.merge_cells('A2:J2')
    
    # Add headers
    headers = [
        'Date', 'Token', 'Order ID', 'Table', 'Order Method', 
        'Items', 'Total Before Tax', 'Tax Amount', 'Total Amount', 'Payment Method'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
    
    # Add data
    row = 5
    total_sales = Decimal('0.00')
    total_tax = Decimal('0.00')
    
    for order in orders:
        # Get items summary
        items_summary = []
        for item in order.items.all():
            items_summary.append(f"{item.quantity}x {item.menu_item.name}")
        items_str = ", ".join(items_summary)
        
        ws.cell(row=row, column=1, value=order.create_date.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=2, value=order.token)
        ws.cell(row=row, column=3, value=order.id)
        ws.cell(row=row, column=4, value=str(order.table) if order.table else order.order_method)
        ws.cell(row=row, column=5, value=order.order_method)
        ws.cell(row=row, column=6, value=items_str)
        ws.cell(row=row, column=7, value=float(order.total_before_tax))
        ws.cell(row=row, column=8, value=float(order.total_tax))
        ws.cell(row=row, column=9, value=float(order.total_price))
        ws.cell(row=row, column=10, value=order.payment_method)
        
        total_sales += order.total_price
        total_tax += order.total_tax
        row += 1
    
    # Add totals
    row += 1
    ws.cell(row=row, column=8, value="TOTALS:").font = header_font
    ws.cell(row=row, column=9, value=float(total_sales)).font = header_font
    ws.cell(row=row, column=8, value=float(total_tax)).font = header_font
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="daybook_report_{start_date}_{end_date}.xlsx"'
    
    wb.save(response)
    return response


def generate_daybook_pdf(orders, start_date, end_date, store):
    """Generate PDF Day Book Report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Add title
    title = Paragraph(f"{store.name} - Day Book Report", title_style)
    subtitle = Paragraph(f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}", styles['Heading2'])
    
    story.append(title)
    story.append(subtitle)
    story.append(Spacer(1, 20))
    
    # Prepare table data
    data = [['Date/Time', 'Token', 'Table', 'Method', 'Items', 'Total', 'Payment']]
    
    total_sales = Decimal('0.00')
    
    for order in orders:
        items_summary = []
        for item in order.items.all()[:2]:  # Limit to 2 items for space
            items_summary.append(f"{item.quantity}x {item.menu_item.name}")
        
        if len(order.items.all()) > 2:
            items_summary.append("...")
            
        items_str = ", ".join(items_summary)
        
        data.append([
            order.create_date.strftime('%m/%d %H:%M'),
            str(order.token),
            str(order.table) if order.table else order.order_method[:8],
            order.order_method[:8],
            items_str[:30],  # Truncate long item lists
            f"${order.total_price:.2f}",
            order.payment_method[:8]
        ])
        
        total_sales += order.total_price
    
    # Add total row
    data.append(['', '', '', '', '', f"Total: ${total_sales:.2f}", ''])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    
    # Build PDF
    doc.build(story)
    
    # Create HTTP response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="daybook_report_{start_date}_{end_date}.pdf"'
    
    return response


@login_required
def generate_sales_summary(request):
    """Generate Sales Summary Report"""
    if request.method == 'POST':
        try:
            store = request.user.store_memberships.all()[0].store
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            format_type = request.POST.get('format', 'excel')
            
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            # Get sales summary data
            orders = Order.objects.filter(
                store=store,
                create_date__date__gte=start_date,
                create_date__date__lte=end_date,
                checkout_status=True,
                payment_status='Paid'
            )
            
            # Group by date
            sales_by_date = orders.extra({'date': "date(create_date)"}).values('date').annotate(
                total_orders=Count('id'),
                total_sales=Sum('total_price'),
                total_tax=Sum('total_tax')
            ).order_by('date')
            
            if format_type == 'excel':
                return generate_sales_summary_excel(sales_by_date, start_date, end_date, store)
            else:
                return generate_sales_summary_pdf(sales_by_date, start_date, end_date, store)
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


def generate_sales_summary_excel(sales_data, start_date, end_date, store):
    """Generate Excel Sales Summary Report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Summary"
    
    # Styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=16)
    
    # Title
    ws['A1'] = f"{store.name} - Sales Summary Report"
    ws['A1'].font = title_font
    ws['A2'] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    ws.merge_cells('A1:E1')
    ws.merge_cells('A2:E2')
    
    # Headers
    headers = ['Date', 'Total Orders', 'Total Sales', 'Total Tax', 'Net Sales']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
    
    # Data
    row = 5
    grand_total_orders = 0
    grand_total_sales = Decimal('0.00')
    grand_total_tax = Decimal('0.00')
    
    for sale in sales_data:
        ws.cell(row=row, column=1, value=sale['date'])
        ws.cell(row=row, column=2, value=sale['total_orders'])
        ws.cell(row=row, column=3, value=float(sale['total_sales'] or 0))
        ws.cell(row=row, column=4, value=float(sale['total_tax'] or 0))
        ws.cell(row=row, column=5, value=float((sale['total_sales'] or 0) - (sale['total_tax'] or 0)))
        
        grand_total_orders += sale['total_orders']
        grand_total_sales += sale['total_sales'] or Decimal('0.00')
        grand_total_tax += sale['total_tax'] or Decimal('0.00')
        row += 1
    
    # Totals
    row += 1
    ws.cell(row=row, column=1, value="TOTALS:").font = header_font
    ws.cell(row=row, column=2, value=grand_total_orders).font = header_font
    ws.cell(row=row, column=3, value=float(grand_total_sales)).font = header_font
    ws.cell(row=row, column=4, value=float(grand_total_tax)).font = header_font
    ws.cell(row=row, column=5, value=float(grand_total_sales - grand_total_tax)).font = header_font
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="sales_summary_{start_date}_{end_date}.xlsx"'
    wb.save(response)
    return response

def generate_order_status_excel(status_data, start_date, end_date, store):
    """Generate Excel Order Status Report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order Status"
    
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=16)
    
    ws['A1'] = f"{store.name} - Order Status Report"
    ws['A1'].font = title_font
    ws['A2'] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    ws.merge_cells('A1:D1')
    ws.merge_cells('A2:D2')
    
    headers = ['Order Status', 'Count', 'Total Value', 'Percentage']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
    
    row = 5
    total_orders = sum([status['count'] for status in status_data])
    
    for status in status_data:
        percentage = (status['count'] / total_orders * 100) if total_orders > 0 else 0
        
        ws.cell(row=row, column=1, value=status['status'])
        ws.cell(row=row, column=2, value=status['count'])
        ws.cell(row=row, column=3, value=float(status['total_value'] or 0))
        ws.cell(row=row, column=4, value=f"{percentage:.2f}%")
        row += 1
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="order_status_{start_date}_{end_date}.xlsx"'
    wb.save(response)
    return response



def generate_sales_summary_pdf(sales_data, start_date, end_date, store):
    """Generate PDF Sales Summary Report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1
    )
    
    title = Paragraph(f"{store.name} - Sales Summary Report", title_style)
    subtitle = Paragraph(f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}", styles['Heading2'])
    
    story.append(title)
    story.append(subtitle)
    story.append(Spacer(1, 20))
    
    # Table data
    data = [['Date', 'Orders', 'Total Sales', 'Tax', 'Net Sales']]
    
    grand_total_orders = 0
    grand_total_sales = Decimal('0.00')
    grand_total_tax = Decimal('0.00')
    
    for sale in sales_data:
        data.append([
            sale['date'].strftime('%Y-%m-%d'),
            str(sale['total_orders']),
            f"${sale['total_sales'] or 0:.2f}",
            f"${sale['total_tax'] or 0:.2f}",
            f"${(sale['total_sales'] or 0) - (sale['total_tax'] or 0):.2f}"
        ])
        
        grand_total_orders += sale['total_orders']
        grand_total_sales += sale['total_sales'] or Decimal('0.00')
        grand_total_tax += sale['total_tax'] or Decimal('0.00')
    
    # Add total row
    data.append([
        'TOTAL',
        str(grand_total_orders),
        f"${grand_total_sales:.2f}",
        f"${grand_total_tax:.2f}",
        f"${grand_total_sales - grand_total_tax:.2f}"
    ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sales_summary_{start_date}_{end_date}.pdf"'
    return response


@login_required
def generate_payment_methods(request):
    """Generate Payment Methods Report"""
    if request.method == 'POST':
        try:
            store = request.user.store_memberships.all()[0].store
            start_date = datetime.strptime(request.POST.get('start_date'), '%Y-%m-%d').date()
            end_date = datetime.strptime(request.POST.get('end_date'), '%Y-%m-%d').date()
            format_type = request.POST.get('format', 'excel')
            
            # Get payment method breakdown
            payment_data = Order.objects.filter(
                store=store,
                create_date__date__gte=start_date,
                create_date__date__lte=end_date,
                checkout_status=True,
                payment_status='Paid'
            ).values('payment_method').annotate(
                count=Count('id'),
                total=Sum('total_price')
            ).order_by('-total')
            
            if format_type == 'excel':
                return generate_payment_methods_excel(payment_data, start_date, end_date, store)
            else:
                return generate_payment_methods_pdf(payment_data, start_date, end_date, store)
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


def generate_payment_methods_excel(payment_data, start_date, end_date, store):
    """Generate Excel Payment Methods Report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Methods"
    
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=16)
    
    ws['A1'] = f"{store.name} - Payment Methods Report"
    ws['A1'].font = title_font
    ws['A2'] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    ws.merge_cells('A1:D1')
    ws.merge_cells('A2:D2')
    
    headers = ['Payment Method', 'Number of Orders', 'Total Amount', 'Percentage']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
    
    row = 5
    total_amount = sum([payment['total'] for payment in payment_data])
    
    for payment in payment_data:
        percentage = (payment['total'] / total_amount * 100) if total_amount > 0 else 0
        
        ws.cell(row=row, column=1, value=payment['payment_method'])
        ws.cell(row=row, column=2, value=payment['count'])
        ws.cell(row=row, column=3, value=float(payment['total']))
        ws.cell(row=row, column=4, value=f"{percentage:.2f}%")
        row += 1
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="payment_methods_{start_date}_{end_date}.xlsx"'
    wb.save(response)
    return response


def generate_payment_methods_pdf(payment_data, start_date, end_date, store):
    """Generate PDF Payment Methods Report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1
    )
    
    title = Paragraph(f"{store.name} - Payment Methods Report", title_style)
    subtitle = Paragraph(f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}", styles['Heading2'])
    
    story.append(title)
    story.append(subtitle)
    story.append(Spacer(1, 20))
    
    # Table data
    data = [['Payment Method', 'Orders', 'Total Amount', 'Percentage']]
    
    total_amount = sum([payment['total'] for payment in payment_data])
    
    for payment in payment_data:
        percentage = (payment['total'] / total_amount * 100) if total_amount > 0 else 0
        data.append([
            payment['payment_method'],
            str(payment['count']),
            f"${payment['total']:.2f}",
            f"{percentage:.2f}%"
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="payment_methods_{start_date}_{end_date}.pdf"'
    return response


@login_required
def generate_menu_performance(request):
    """Generate Menu Items Performance Report"""
    if request.method == 'POST':
        try:
            store = request.user.store_memberships.all()[0].store
            start_date = datetime.strptime(request.POST.get('start_date'), '%Y-%m-%d').date()
            end_date = datetime.strptime(request.POST.get('end_date'), '%Y-%m-%d').date()
            format_type = request.POST.get('format', 'excel')
            
            # Get menu performance data
            menu_data = OrderItem.objects.filter(
                order__store=store,
                order__create_date__date__gte=start_date,
                order__create_date__date__lte=end_date,
                order__checkout_status=True,
                is_saved_for_later=False
            ).values('menu_item__name', 'menu_item__category__name').annotate(
                total_quantity=Sum('quantity'),
                total_revenue=Sum(F('price') * F('quantity'))
            ).order_by('-total_quantity')
            
            if format_type == 'excel':
                return generate_menu_performance_excel(menu_data, start_date, end_date, store)
            else:
                return generate_menu_performance_pdf(menu_data, start_date, end_date, store)
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


def generate_menu_performance_excel(menu_data, start_date, end_date, store):
    """Generate Excel Menu Performance Report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Menu Performance"
    
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=16)
    
    ws['A1'] = f"{store.name} - Menu Performance Report"
    ws['A1'].font = title_font
    ws['A2'] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    ws.merge_cells('A1:E1')
    ws.merge_cells('A2:E2')
    
    headers = ['Menu Item', 'Category', 'Quantity Sold', 'Revenue', 'Avg Price']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
    
    row = 5
    for item in menu_data:
        avg_price = item['total_revenue'] / item['total_quantity'] if item['total_quantity'] > 0 else 0
        
        ws.cell(row=row, column=1, value=item['menu_item__name'])
        ws.cell(row=row, column=2, value=item['menu_item__category__name'])
        ws.cell(row=row, column=3, value=item['total_quantity'])
        ws.cell(row=row, column=4, value=float(item['total_revenue']))
        ws.cell(row=row, column=5, value=f"{avg_price:.2f}")
        row += 1
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="menu_performance_{start_date}_{end_date}.xlsx"'
    wb.save(response)
    return response


@login_required
def generate_tax_report(request):
    """Generate Tax Report"""
    if request.method == 'POST':
        try:
            store = request.user.store_memberships.all()[0].store
            start_date = datetime.strptime(request.POST.get('start_date'), '%Y-%m-%d').date()
            end_date = datetime.strptime(request.POST.get('end_date'), '%Y-%m-%d').date()
            format_type = request.POST.get('format', 'excel')
            
            # Get tax breakdown
            orders = Order.objects.filter(
                store=store,
                create_date__date__gte=start_date,
                create_date__date__lte=end_date,
                checkout_status=True,
                payment_status='Paid'
            )
            
            tax_summary = {
                'total_before_tax': orders.aggregate(Sum('total_before_tax'))['total_before_tax__sum'] or Decimal('0'),
                'total_tax': orders.aggregate(Sum('total_tax'))['total_tax__sum'] or Decimal('0'),
                'total_with_tax': orders.aggregate(Sum('total_price'))['total_price__sum'] or Decimal('0'),
                'order_count': orders.count()
            }
            
            if format_type == 'excel':
                return generate_tax_report_excel(tax_summary, orders, start_date, end_date, store)
            else:
                return generate_tax_report_pdf(tax_summary, orders, start_date, end_date, store)
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


def generate_tax_report_excel(tax_summary, orders, start_date, end_date, store):
    """Generate Excel Tax Report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tax Report"
    
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=16)
    
    ws['A1'] = f"{store.name} - Tax Report"
    ws['A1'].font = title_font
    ws['A2'] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    ws.merge_cells('A1:C1')
    ws.merge_cells('A2:C2')
    
    # Summary section
    ws['A4'] = "TAX SUMMARY"
    ws['A4'].font = header_font
    
    ws['A5'] = "Total Orders:"
    ws['B5'] = tax_summary['order_count']
    
    ws['A6'] = "Total Before Tax:"
    ws['B6'] = float(tax_summary['total_before_tax'])
    
    ws['A7'] = "Total Tax Amount:"
    ws['B7'] = float(tax_summary['total_tax'])
    
    ws['A8'] = "Total With Tax:"
    ws['B8'] = float(tax_summary['total_with_tax'])
    
    # Calculate effective tax rate
    if tax_summary['total_before_tax'] > 0:
        tax_rate = (tax_summary['total_tax'] / tax_summary['total_before_tax']) * 100
        ws['A9'] = "Effective Tax Rate:"
        ws['B9'] = f"{tax_rate:.2f}%"
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="tax_report_{start_date}_{end_date}.xlsx"'
    wb.save(response)
    return response


@login_required
def generate_order_status(request):
    """Generate Order Status Report"""
    if request.method == 'POST':
        try:
            store = request.user.store_memberships.all()[0].store
            start_date = datetime.strptime(request.POST.get('start_date'), '%Y-%m-%d').date()
            end_date = datetime.strptime(request.POST.get('end_date'), '%Y-%m-%d').date()
            format_type = request.POST.get('format', 'excel')
            
            # Get order status breakdown
            status_data = Order.objects.filter(
                store=store,
                create_date__date__gte=start_date,
                create_date__date__lte=end_date
            ).values('status').annotate(
                count=Count('id'),
                total_value=Sum('total_price')
            ).order_by('-count')
            
            if format_type == 'excel':
                return generate_order_status_excel(status_data, start_date, end_date, store)
            else:
                return generate_order_status_pdf(status_data, start_date, end_date, store)
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


def generate_menu_performance_pdf(menu_data, start_date, end_date, store):
    """Generate PDF Menu Performance Report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1
    )
    
    title = Paragraph(f"{store.name} - Menu Performance Report", title_style)
    subtitle = Paragraph(f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}", styles['Heading2'])
    
    story.append(title)
    story.append(subtitle)
    story.append(Spacer(1, 20))
    
    # Table data
    data = [['Menu Item', 'Category', 'Qty Sold', 'Revenue', 'Avg Price']]
    
    for item in menu_data:
        avg_price = item['total_revenue'] / item['total_quantity'] if item['total_quantity'] > 0 else 0
        
        data.append([
            item['menu_item__name'][:25],  # Truncate long names
            item['menu_item__category__name'][:15],
            str(item['total_quantity']),
            f"${item['total_revenue']:.2f}",
            f"${avg_price:.2f}"
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),  # Smaller font for data
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="menu_performance_{start_date}_{end_date}.pdf"'
    return response


def generate_tax_report_pdf(tax_summary, orders, start_date, end_date, store):
    """Generate PDF Tax Report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1
    )
    
    title = Paragraph(f"{store.name} - Tax Report", title_style)
    subtitle = Paragraph(f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}", styles['Heading2'])
    
    story.append(title)
    story.append(subtitle)
    story.append(Spacer(1, 20))
    
    # Summary table
    summary_data = [
        ['Tax Summary', 'Amount'],
        ['Total Orders', str(tax_summary['order_count'])],
        ['Total Before Tax', f"${tax_summary['total_before_tax']:.2f}"],
        ['Total Tax Amount', f"${tax_summary['total_tax']:.2f}"],
        ['Total With Tax', f"${tax_summary['total_with_tax']:.2f}"],
    ]
    
    # Calculate tax percentage
    if tax_summary['total_before_tax'] > 0:
        tax_percentage = (tax_summary['total_tax'] / tax_summary['total_before_tax']) * 100
        summary_data.append(['Effective Tax Rate', f"{tax_percentage:.2f}%"])
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),  # Bold total row
    ]))
    
    story.append(summary_table)
    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="tax_report_{start_date}_{end_date}.pdf"'
    return response


def generate_order_status_pdf(status_data, start_date, end_date, store):
    """Generate PDF Order Status Report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1
    )
    
    title = Paragraph(f"{store.name} - Order Status Report", title_style)
    subtitle = Paragraph(f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}", styles['Heading2'])
    
    story.append(title)
    story.append(subtitle)
    story.append(Spacer(1, 20))
    
    # Table data
    data = [['Order Status', 'Count', 'Total Value', 'Percentage']]
    
    total_orders = sum([status['count'] for status in status_data])
    total_value = sum([status['total_value'] or 0 for status in status_data])
    
    for status in status_data:
        percentage = (status['count'] / total_orders * 100) if total_orders > 0 else 0
        
        data.append([
            status['status'],
            str(status['count']),
            f"${status['total_value'] or 0:.2f}",
            f"{percentage:.2f}%"
        ])
    
    # Add total row
    data.append([
        'TOTAL',
        str(total_orders),
        f"${total_value:.2f}",
        '100.00%'
    ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="order_status_{start_date}_{end_date}.pdf"'
    return response


from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from decimal import Decimal
from datetime import datetime
import json
import traceback

from inventory.models import Menu, Tax, FoodCategory, Modifiers
from orders.models import Order, OrderItem, Tables, Checkout
from authentication.models import Store


@login_required
def b2b_pos(request):
    """Main B2B POS Screen"""
    try:
        # Get user's store
        store_membership = request.user.store_memberships.first()
        if not store_membership:
            return render(request, 'error.html', {'message': 'No store assigned'})
        
        store = store_membership.store
        
        # Get all menu items for this store
        menu_items = Menu.objects.filter(
            store=store,
            status=True
        ).select_related('category').prefetch_related('taxes', 'modifiers')
        
        # Get categories
        categories = FoodCategory.objects.filter(store=store, active=True)
        
        # Get tables for order method
        tables = Tables.objects.filter(status=True).order_by('Table_number')
        
        # Get taxes
        taxes = Tax.objects.filter(store=store, is_active=True)
        
        # Get active B2B orders (not completed)
        active_orders = Order.objects.filter(
            store=store,
            order_method='B2B',
            completion_status=False
        ).order_by('-create_date')[:10]
        
        context = {
            'menu_items': menu_items,
            'categories': categories,
            'tables': tables,
            'taxes': taxes,
            'active_orders': active_orders,
            'store': store,
        }
        
        return render(request, 'b2b/index.html', context)
    
    except Exception as e:
        print(f"Error in b2b_pos: {str(e)}")
        traceback.print_exc()
        return render(request, 'error.html', {'message': str(e)})


@login_required
@require_http_methods(["POST"])
def create_b2b_order(request):
    """Create a new B2B order"""
    try:
        # Try to parse JSON, if it fails, it might be empty POST
        try:
            data = json.loads(request.body) if request.body else {}
        except json.JSONDecodeError:
            data = {}
        
        store_membership = request.user.store_memberships.first()
        
        if not store_membership:
            return JsonResponse({
                'success': False, 
                'message': 'No store assigned to user'
            }, status=400)
        
        store = store_membership.store
        
        with transaction.atomic():
            # Create order
            order = Order.objects.create(
                store=store,
                user=request.user,
                order_method='B2B',
                status='Pending',
                checkout_status=False,
                completion_status=False
            )
            
            return JsonResponse({
                'success': True,
                'order_id': order.id,
                'token': order.token,
                'message': 'B2B Order created successfully'
            })
    
    except Exception as e:
        print(f"Error in create_b2b_order: {str(e)}")
        traceback.print_exc()
        return JsonResponse({
            'success': False, 
            'message': f'Error creating order: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def add_b2b_item(request):
    """Add item to B2B order with custom price"""
    try:
        # Parse JSON data
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError as e:
            return JsonResponse({
                'success': False, 
                'message': f'Invalid JSON data: {str(e)}'
            }, status=400)
        
        order_id = data.get('order_id')
        menu_item_id = data.get('menu_item_id')
        custom_price = data.get('custom_price')
        quantity = data.get('quantity', 1)
        special_instructions = data.get('special_instructions', '')
        addon_ids = data.get('addon_ids', [])
        
        # Validate required fields
        if not order_id:
            return JsonResponse({
                'success': False, 
                'message': 'Order ID is required'
            }, status=400)
        
        if not menu_item_id:
            return JsonResponse({
                'success': False, 
                'message': 'Menu item ID is required'
            }, status=400)
        
        if not custom_price:
            return JsonResponse({
                'success': False, 
                'message': 'Custom price is required'
            }, status=400)
        
        # Convert to proper types
        try:
            custom_price = Decimal(str(custom_price))
            quantity = int(quantity)
        except (ValueError, TypeError) as e:
            return JsonResponse({
                'success': False, 
                'message': f'Invalid price or quantity: {str(e)}'
            }, status=400)
        
        # Validate custom price
        if custom_price <= 0:
            return JsonResponse({
                'success': False, 
                'message': 'Price must be greater than 0'
            }, status=400)
        
        # Get objects
        try:
            order = Order.objects.get(id=order_id)
            menu_item = Menu.objects.get(id=menu_item_id)
        except Order.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'message': 'Order not found'
            }, status=404)
        except Menu.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'message': 'Menu item not found'
            }, status=404)
        
        with transaction.atomic():
            # Create order item with custom price
            order_item = OrderItem.objects.create(
                order=order,
                menu_item=menu_item,
                quantity=quantity,
                price=custom_price,
                special_instructions=special_instructions,
                is_saved_for_later=False
            )
            
            # Add taxes from menu item
            if menu_item.taxes.exists():
                order_item.tax.set(menu_item.taxes.all())
            
            # Add modifiers/addons if selected
            if addon_ids:
                modifiers = Modifiers.objects.filter(id__in=addon_ids)
                order_item.add_ons.set(modifiers)
            
            # Calculate totals
            order.calculate_totals()
            order.save()
            
            # Get item details for response
            addon_total = sum(addon.price for addon in order_item.add_ons.all())
            item_data = {
                'id': order_item.id,
                'name': menu_item.name,
                'quantity': quantity,
                'price': float(custom_price),
                'addon_total': float(addon_total),
                'total': float(order_item.get_total_price_with_addons()),
                'tax_amount': float(order_item.get_tax_amount())
            }
            
            return JsonResponse({
                'success': True,
                'message': 'Item added successfully',
                'item': item_data,
                'order_total': float(order.total_price),
                'order_tax': float(order.total_tax),
                'order_subtotal': float(order.total_before_tax)
            })
    
    except Exception as e:
        print(f"Error in add_b2b_item: {str(e)}")
        traceback.print_exc()
        return JsonResponse({
            'success': False, 
            'message': f'Error adding item: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def update_b2b_item(request):
    """Update B2B order item (quantity or price)"""
    try:
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError as e:
            return JsonResponse({
                'success': False, 
                'message': f'Invalid JSON data: {str(e)}'
            }, status=400)
        
        item_id = data.get('item_id')
        quantity = data.get('quantity')
        custom_price = data.get('custom_price')
        
        if not item_id:
            return JsonResponse({
                'success': False, 
                'message': 'Item ID is required'
            }, status=400)
        
        try:
            order_item = OrderItem.objects.get(id=item_id)
        except OrderItem.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'message': 'Order item not found'
            }, status=404)
        
        with transaction.atomic():
            if quantity is not None:
                try:
                    order_item.quantity = int(quantity)
                except (ValueError, TypeError):
                    return JsonResponse({
                        'success': False, 
                        'message': 'Invalid quantity'
                    }, status=400)
            
            if custom_price is not None:
                try:
                    order_item.price = Decimal(str(custom_price))
                except (ValueError, TypeError):
                    return JsonResponse({
                        'success': False, 
                        'message': 'Invalid price'
                    }, status=400)
            
            order_item.save()
            
            # Recalculate order totals
            order_item.order.calculate_totals()
            order_item.order.save()
            
            addon_total = sum(addon.price for addon in order_item.add_ons.all())
            
            return JsonResponse({
                'success': True,
                'message': 'Item updated successfully',
                'item_total': float(order_item.get_total_price_with_addons()),
                'order_total': float(order_item.order.total_price),
                'order_tax': float(order_item.order.total_tax),
                'order_subtotal': float(order_item.order.total_before_tax)
            })
    
    except Exception as e:
        print(f"Error in update_b2b_item: {str(e)}")
        traceback.print_exc()
        return JsonResponse({
            'success': False, 
            'message': f'Error updating item: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def remove_b2b_item(request):
    """Remove item from B2B order"""
    try:
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError as e:
            return JsonResponse({
                'success': False, 
                'message': f'Invalid JSON data: {str(e)}'
            }, status=400)
        
        item_id = data.get('item_id')
        
        if not item_id:
            return JsonResponse({
                'success': False, 
                'message': 'Item ID is required'
            }, status=400)
        
        try:
            order_item = OrderItem.objects.get(id=item_id)
        except OrderItem.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'message': 'Order item not found'
            }, status=404)
        
        order = order_item.order
        
        with transaction.atomic():
            order_item.delete()
            
            # Recalculate order totals
            order.calculate_totals()
            order.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Item removed successfully',
                'order_total': float(order.total_price),
                'order_tax': float(order.total_tax),
                'order_subtotal': float(order.total_before_tax)
            })
    
    except Exception as e:
        print(f"Error in remove_b2b_item: {str(e)}")
        traceback.print_exc()
        return JsonResponse({
            'success': False, 
            'message': f'Error removing item: {str(e)}'
        }, status=500)


# @login_required
# @require_http_methods(["GET"])
# def get_b2b_order_details(request, order_id):
#     """Get B2B order details"""
#     try:
#         try:
#             order = Order.objects.get(id=order_id)
#         except Order.DoesNotExist:
#             return JsonResponse({
#                 'success': False, 
#                 'message': 'Order not found'
#             }, status=404)
        
#         items = []
#         for item in order.items.all():
#             addon_total = sum(Decimal(str(addon.price)) for addon in item.add_ons.all())
#             items.append({
#                 'id': item.id,
#                 'menu_item_id': item.menu_item.id,
#                 'name': item.menu_item.name,
#                 'quantity': item.quantity,
#                 'price': float(item.price),
#                 'addon_total': float(addon_total),
#                 'total': float(item.get_total_price_with_addons()),
#                 'tax_amount': float(item.get_tax_amount()),
#                 'special_instructions': item.special_instructions or '',
#                 'addons': [{'id': a.id, 'name': a.name, 'price': float(a.price)} 
#                           for a in item.add_ons.all()]
#             })
        
#         return JsonResponse({
#             'success': True,
#             'order': {
#                 'id': order.id,
#                 'token': order.token,
#                 'status': order.status,
#                 'order_method': order.order_method,
#                 'total_price': float(order.total_price),
#                 'total_tax': float(order.total_tax),
#                 'total_before_tax': float(order.total_before_tax),
#                 'items': items
#             }
#         })
    
#     except Exception as e:
#         print(f"Error in get_b2b_order_details: {str(e)}")
#         traceback.print_exc()
#         return JsonResponse({
#             'success': False, 
#             'message': f'Error getting order details: {str(e)}'
#         }, status=500)


@login_required
@require_http_methods(["POST"])
def checkout_b2b_order(request):
    """Checkout B2B order"""
    try:
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError as e:
            return JsonResponse({
                'success': False, 
                'message': f'Invalid JSON data: {str(e)}'
            }, status=400)
        
        order_id = data.get('order_id')
        payment_method = data.get('payment_method', 'Cash')
        customer_name = data.get('customer_name', '')
        customer_phone = data.get('customer_phone', '')
        discount_amount = Decimal(str(data.get('discount_amount', 0)))
        discount_reason = data.get('discount_reason', '')
        
        if not order_id:
            return JsonResponse({
                'success': False, 
                'message': 'Order ID is required'
            }, status=400)
        
        try:
            order = Order.objects.get(id=order_id)
        except Order.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'message': 'Order not found'
            }, status=404)
        
        with transaction.atomic():
            # Create checkout
            checkout = Checkout.objects.create(
                order=order,
                total_price=order.total_price,
                tax_amount=order.total_tax,
                payment_method=payment_method,
                payment_status='Paid',
                customer_name=customer_name,
                customer_phone=customer_phone,
                discount_amount=discount_amount,
                discount_reason=discount_reason
            )
            
            # Update order status
            order.checkout_status = True
            order.status = 'Completed'
            order.completion_status = True
            order.payment_method = payment_method
            order.payment_status = 'Paid'
            order.save()
            
            final_amount = checkout.calculate_final_amount()
            
            return JsonResponse({
                'success': True,
                'message': 'Order completed successfully',
                'checkout_id': checkout.id,
                'final_amount': float(final_amount),
                'order_token': order.token
            })
    
    except Exception as e:
        print(f"Error in checkout_b2b_order: {str(e)}")
        traceback.print_exc()
        return JsonResponse({
            'success': False, 
            'message': f'Error completing checkout: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def search_menu_items(request):
    """Search menu items for B2B POS"""
    try:
        query = request.GET.get('q', '')
        store_membership = request.user.store_memberships.first()
        
        if not store_membership:
            return JsonResponse({
                'success': False, 
                'message': 'No store assigned'
            }, status=400)
        
        store = store_membership.store
        
        menu_items = Menu.objects.filter(
            store=store,
            status=True
        ).filter(
            name__icontains=query
        ) | Menu.objects.filter(
            store=store,
            status=True,
            code__icontains=query
        ) | Menu.objects.filter(
            store=store,
            status=True,
            barcode__icontains=query
        )
        
        results = []
        for item in menu_items[:20]:  # Limit to 20 results
            results.append({
                'id': item.id,
                'name': item.name,
                'category': item.category.name,
                'price': float(item.price),
                'code': item.code or '',
                'image': item.image.url if item.image else None
            })
        
        return JsonResponse({'success': True, 'items': results})
    
    except Exception as e:
        print(f"Error in search_menu_items: {str(e)}")
        traceback.print_exc()
        return JsonResponse({
            'success': False, 
            'message': f'Error searching items: {str(e)}'
        }, status=500)
    

@login_required
@require_http_methods(["GET"])
def get_active_b2b_orders(request):
    """Get active B2B orders for live updates"""
    try:
        store_membership = request.user.store_memberships.first()
        
        if not store_membership:
            return JsonResponse({
                'success': False, 
                'message': 'No store assigned'
            }, status=400)
        
        store = store_membership.store
        
        # Get active B2B orders
        active_orders = Order.objects.filter(
            store=store,
            order_method='B2B',
            completion_status=False
        ).order_by('-create_date')[:10]
        
        orders_data = []
        for order in active_orders:
            orders_data.append({
                'id': order.id,
                'token': order.token,
                'status': order.status,
                'total_price': float(order.total_price or 0),
                'time': order.create_date.strftime('%I:%M %p'),
                'items_count': order.items.filter(is_saved_for_later=False).count()
            })
        
        return JsonResponse({
            'success': True,
            'orders': orders_data
        })
    
    except Exception as e:
        print(f"Error in get_active_b2b_orders: {str(e)}")
        traceback.print_exc()
        return JsonResponse({
            'success': False, 
            'message': f'Error getting orders: {str(e)}'
        }, status=500)
# Make sure you also have this view (it should already exist in your views.py)

@login_required
@require_http_methods(["GET"])
def get_b2b_order_details(request, order_id):
    """Get B2B order details"""
    try:
        try:
            order = Order.objects.get(id=order_id)
        except Order.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'message': 'Order not found'
            }, status=404)
        
        # Only get items not saved for later
        items = []
        order_items = order.items.filter(is_saved_for_later=False)
        
        for item in order_items:
            addon_total = sum(Decimal(str(addon.price)) for addon in item.add_ons.all())
            
            # Calculate item total with addons
            item_base_total = (Decimal(str(item.price)) + addon_total) * item.quantity
            
            # Calculate tax for this item
            item_tax_total = Decimal('0.00')
            for tax in item.tax.all():
                tax_amount = (item_base_total * tax.tax_percentage) / 100
                item_tax_total += tax_amount
            
            items.append({
                'id': item.id,
                'menu_item_id': item.menu_item.id,
                'name': item.menu_item.name,
                'quantity': item.quantity,
                'price': float(item.price),
                'addon_total': float(addon_total),
                'total': float(item_base_total),
                'tax_amount': float(item_tax_total),
                'special_instructions': item.special_instructions or '',
                'addons': [{'id': a.id, 'name': a.name, 'price': float(a.price)} 
                          for a in item.add_ons.all()]
            })
        
        return JsonResponse({
            'success': True,
            'order': {
                'id': order.id,
                'token': order.token,
                'status': order.status,
                'order_method': order.order_method,
                'total_price': float(order.total_price or 0),
                'total_tax': float(order.total_tax or 0),
                'total_before_tax': float(order.total_before_tax or 0),
                'items': items
            }
        })
    
    except Exception as e:
        print(f"Error in get_b2b_order_details: {str(e)}")
        traceback.print_exc()
        return JsonResponse({
            'success': False, 
            'message': f'Error getting order details: {str(e)}'
        }, status=500)

# b2b sale list 

@login_required
def b2b_sales_list(request):
    """B2B Sales List with filtering and sorting"""
    try:
        store_membership = request.user.store_memberships.first()
        if not store_membership:
            return render(request, 'error.html', {'message': 'No store assigned'})
        
        store = store_membership.store
        
        # Base queryset - ALL B2B orders (including pending)
        orders = Order.objects.filter(
            store=store,
            order_method='B2B'
        ).select_related('checkout').prefetch_related('items')
        
        # Get filter parameters
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        status = request.GET.get('status')
        payment_status = request.GET.get('payment_status')
        payment_method = request.GET.get('payment_method')
        completion = request.GET.get('completion')
        search = request.GET.get('search')
        
        # Apply filters
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                orders = orders.filter(create_date__gte=date_from_obj)
            except ValueError:
                pass
        
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                # Add one day to include the entire end date
                date_to_obj = date_to_obj + timedelta(days=1)
                orders = orders.filter(create_date__lt=date_to_obj)
            except ValueError:
                pass
        
        if status:
            orders = orders.filter(status=status)
        
        if payment_status:
            orders = orders.filter(payment_status=payment_status)
        
        if payment_method:
            orders = orders.filter(payment_method=payment_method)
        
        if completion:
            if completion == 'completed':
                orders = orders.filter(completion_status=True)
            elif completion == 'pending':
                orders = orders.filter(completion_status=False)
        
        if search:
            orders = orders.filter(
                Q(token__icontains=search) |
                Q(checkout__customer_name__icontains=search) |
                Q(checkout__customer_phone__icontains=search)
            )
        
        # Sorting
        sort_by = request.GET.get('sort', 'create_date')
        order_direction = request.GET.get('order', 'desc')
        
        valid_sort_fields = ['token', 'create_date', 'total_price', 'status']
        if sort_by not in valid_sort_fields:
            sort_by = 'create_date'
        
        if order_direction == 'asc':
            orders = orders.order_by(sort_by)
        else:
            orders = orders.order_by(f'-{sort_by}')
        
        # Calculate statistics
        stats = {
            'total_orders': orders.count(),
            'total_sales': orders.aggregate(total=Sum('total_price'))['total'] or Decimal('0.00'),
            'total_tax': orders.aggregate(tax=Sum('total_tax'))['tax'] or Decimal('0.00'),
            'avg_order_value': orders.aggregate(avg=Avg('total_price'))['avg'] or Decimal('0.00'),
        }
        
        # Completion stats
        stats['completed_orders'] = orders.filter(completion_status=True).count()
        stats['pending_orders'] = orders.filter(completion_status=False).count()
        
        # Payment stats
        stats['paid_orders'] = orders.filter(payment_status='Paid').count()
        stats['pending_payment'] = orders.filter(payment_status='Pending').count()
        
        # Today's stats
        today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
        today_orders = orders.filter(create_date__gte=today_start)
        stats['today_orders'] = today_orders.count()
        stats['today_sales'] = today_orders.aggregate(total=Sum('total_price'))['total'] or Decimal('0.00')
        
        # Check for export request
        export_type = request.GET.get('export')
        if export_type == 'excel':
            return export_b2b_sales_excel(orders, stats)
        elif export_type == 'pdf':
            return export_b2b_sales_pdf(orders, stats)
        
        # Pagination
        paginator = Paginator(orders, 25)  # 25 orders per page
        page_number = request.GET.get('page', 1)
        
        try:
            page_obj = paginator.get_page(page_number)
        except:
            page_obj = paginator.get_page(1)
        
        # Build query string for pagination links
        query_params = request.GET.copy()
        if 'page' in query_params:
            query_params.pop('page')
        query_string = query_params.urlencode()
        
        context = {
            'orders': page_obj,
            'stats': stats,
            'filters': {
                'date_from': date_from or '',
                'date_to': date_to or '',
                'status': status or '',
                'payment_status': payment_status or '',
                'payment_method': payment_method or '',
                'completion': completion or '',
                'search': search or '',
            },
            'sort_by': sort_by,
            'order': order_direction,
            'query_string': query_string,
        }
        
        return render(request, 'b2b/sales_list.html', context)
    
    except Exception as e:
        print(f"Error in b2b_sales_list: {str(e)}")
        traceback.print_exc()
        return render(request, 'error.html', {'message': str(e)})


def export_b2b_sales_excel(orders, stats):
    """Export B2B sales to Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "B2B Sales Report"
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Add summary
    ws['A1'] = 'B2B SALES REPORT'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:I1')
    
    ws['A3'] = 'Total Orders:'
    ws['B3'] = stats['total_orders']
    ws['A4'] = 'Total Sales:'
    ws['B4'] = float(stats['total_sales'])
    ws['A5'] = 'Total Tax:'
    ws['B5'] = float(stats['total_tax'])
    ws['A6'] = 'Average Order:'
    ws['B6'] = float(stats['avg_order_value'])
    
    # Headers
    headers = ['Token', 'Date', 'Time', 'Customer', 'Phone', 'Items', 'Payment', 'Amount', 'Tax', 'Status']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = 9
    for order in orders:
        ws.cell(row=row, column=1).value = f"#{order.token}"
        ws.cell(row=row, column=2).value = order.create_date.strftime('%Y-%m-%d')
        ws.cell(row=row, column=3).value = order.create_date.strftime('%H:%M:%S')
        
        if hasattr(order, 'checkout') and order.checkout:
            ws.cell(row=row, column=4).value = order.checkout.customer_name or 'Walk-in'
            ws.cell(row=row, column=5).value = order.checkout.customer_phone or ''
        else:
            ws.cell(row=row, column=4).value = 'Walk-in'
            ws.cell(row=row, column=5).value = ''
        
        ws.cell(row=row, column=6).value = order.items.count()
        ws.cell(row=row, column=7).value = order.payment_method or 'N/A'
        ws.cell(row=row, column=8).value = float(order.total_price)
        ws.cell(row=row, column=9).value = float(order.total_tax)
        ws.cell(row=row, column=10).value = order.status
        
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=B2B_Sales_Report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


def export_b2b_sales_pdf(orders, stats):
    """Export B2B sales to PDF"""
    from django.http import HttpResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    import io
    
    # Create buffer
    buffer = io.BytesIO()
    
    # Create PDF
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#4472C4'),
        alignment=TA_CENTER,
        spaceAfter=30
    )
    
    # Title
    elements.append(Paragraph("B2B SALES REPORT", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary statistics
    summary_data = [
        ['Total Orders', 'Total Sales', 'Total Tax', 'Average Order'],
        [
            str(stats['total_orders']),
            f"₹{stats['total_sales']:.2f}",
            f"₹{stats['total_tax']:.2f}",
            f"₹{stats['avg_order_value']:.2f}"
        ]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch, 2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Orders table
    table_data = [['Token', 'Date', 'Customer', 'Items', 'Payment', 'Amount', 'Tax', 'Status']]
    
    for order in orders:
        customer = 'Walk-in'
        if hasattr(order, 'checkout') and order.checkout and order.checkout.customer_name:
            customer = order.checkout.customer_name
        
        table_data.append([
            f"#{order.token}",
            order.create_date.strftime('%d/%m/%Y %H:%M'),
            customer[:20],  # Truncate long names
            str(order.items.count()),
            order.payment_method or 'N/A',
            f"₹{order.total_price:.2f}",
            f"₹{order.total_tax:.2f}",
            order.status
        ])
    
    orders_table = Table(table_data, colWidths=[0.8*inch, 1.3*inch, 1.5*inch, 0.7*inch, 1*inch, 1*inch, 0.9*inch, 1*inch])
    orders_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(orders_table)
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF value
    pdf = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=B2B_Sales_Report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response.write(pdf)
    
    return response


@login_required
def b2b_invoice(request, order_id):
    """Generate printable B2B invoice"""
    try:
        order = Order.objects.get(id=order_id)
        store_membership = request.user.store_memberships.first()
        
        if not store_membership:
            return render(request, 'error.html', {'message': 'No store assigned'})
        
        store = store_membership.store
        
        # Get order items
        items = order.items.filter(is_saved_for_later=False)
        
        context = {
            'order': order,
            'store': store,
            'items': items,
        }
        
        return render(request, 'b2b/invoice.html', context)
    
    except Order.DoesNotExist:
        return render(request, 'error.html', {'message': 'Order not found'})
    except Exception as e:
        print(f"Error in b2b_invoice: {str(e)}")
        traceback.print_exc()
        return render(request, 'error.html', {'message': str(e)})